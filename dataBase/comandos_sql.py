import sqlite3
# from constants import *
from openpyxl import load_workbook
POL2M = 0.0254

def conect_sqlite(data_base, query = None):
    banco = sqlite3.connect(data_base)
    cursor = banco.cursor()
    if query != None:
        cursor.execute(query)
    banco.commit()
    return cursor

def filtro_sqlite(cursor, query, one = False):
   
    cursor.execute(query)
    #print(cursor.fetchall())
    #cursor.commit()
    if one:
        result = cursor.fetchone()
    else:
        result = cursor.fetchall()
    return result


if __name__ == "__main__":
    # conect_sqlite(DB_CONSTANTS_DIR, "CREATE TABLE Contagem_de_tubos(Ds DOUBLE, Dotl DOUBLE, de_pol DOUBLE, arranjo TEXT, p_pol DOUBLE, Npt1 INTEGER, Npt2 INTEGER, Npt4 INTEGER, Npt6 INTEGER, Npt8 INTEGER)")
    # conect_sqlite(DB_CONSTANTS_DIR, "DROP TABLE Contagem_de_tubos") 
    
    wb = load_workbook(r"C:\Users\carvalhoe\Documents\GITHUB\HeatExGA\base_dados.xlsx")
    sheet = wb["SQL"]

    for row in sheet.iter_rows(values_only=True):
        Ds = row[0] * POL2M
        Dotl = row[1]  *POL2M
        de = row[2]
        arranjo_num = row[3]

        Np1 = row[4]
        Np2 = row[5]
        Np4 = row[6]
        Np6 = row[7]
        Np8 = row[8]

        if Np8 == None:
            Np8 = 0
        print(arranjo_num)
        if arranjo_num ==1:
            p = 15/16
            arranjo = "triangular"
            sql = f"INSERT INTO Contagem_de_tubos VALUES ({Ds}, {Dotl}, {de}, '{arranjo}', {p}, {Np1}, {Np2}, {Np4}, {Np6}, {Np8});"

            conect_sqlite(DB_CONSTANTS_DIR, sql)
        elif arranjo_num ==2:
            p =1
            arranjo = "rodado"
            sql = f"INSERT INTO Contagem_de_tubos VALUES ({Ds}, {Dotl}, {de}, '{arranjo}', {p}, {Np1}, {Np2}, {Np4}, {Np6}, {Np8});"

            conect_sqlite(DB_CONSTANTS_DIR, sql)
            
            p = 1
            arranjo = "quadrado"
            sql = f"INSERT INTO Contagem_de_tubos VALUES ({Ds}, {Dotl}, {de}, '{arranjo}', {p}, {Np1}, {Np2}, {Np4}, {Np6}, {Np8});"
            conect_sqlite(DB_CONSTANTS_DIR, sql)
        elif arranjo_num == 3:
            p=1
            arranjo = "triangular"
            sql = f"INSERT INTO Contagem_de_tubos VALUES ({Ds}, {Dotl}, {de}, '{arranjo}', {p}, {Np1}, {Np2}, {Np4}, {Np6}, {Np8});"
            conect_sqlite(DB_CONSTANTS_DIR, sql)        
        elif arranjo_num == 4:
            p =1 + 1/4
            arranjo = "rodado"
            sql = f"INSERT INTO Contagem_de_tubos VALUES ({Ds}, {Dotl}, {de}, '{arranjo}', {p}, {Np1}, {Np2}, {Np4}, {Np6}, {Np8});"
            conect_sqlite(DB_CONSTANTS_DIR, sql)

            p = 1 + 1/4
            arranjo = "quadrado"
            sql = f"INSERT INTO Contagem_de_tubos VALUES ({Ds}, {Dotl}, {de}, '{arranjo}', {p}, {Np1}, {Np2}, {Np4}, {Np6}, {Np8});"
            conect_sqlite(DB_CONSTANTS_DIR, sql)            
        elif arranjo_num == 5:
            p =1 + 1/4
            arranjo = "triangular"
            sql = f"INSERT INTO Contagem_de_tubos VALUES ({Ds}, {Dotl}, {de}, '{arranjo}', {p}, {Np1}, {Np2}, {Np4}, {Np6}, {Np8});"
            conect_sqlite(DB_CONSTANTS_DIR, sql)
            


    # de = 0.750 * POL2M
    # p = 1 * POL2M
    # a_tubos = "triangular"
    # pp = 0.866 * POL2M
    # pn = 0.884 * POL2M

    # npt = "NP_1"
    # a_tubos = 'triangular 1 pol'
    # Ds = 0.254
    # de = 0.01905
    # cursor = conect_sqlite(DB_CONSTANTS_DIR)
    # sql_NT = f"SELECT {npt} FROM Contagem_de_tubos WHERE a_tubos = '{a_tubos}' AND Ds_m = {Ds} AND d_m = {de}"
    # sql_Dotl = f"SELECT Dotl_m FROM Contagem_de_tubos WHERE a_tubos = '{a_tubos}' AND Ds_m = {Ds} AND d_m = {de}"

    # Nt = filtro_sqlite(cursor, sql_NT, True)
    # Dotl = filtro_sqlite(cursor, sql_Dotl, True)

    # print(Nt)
    # print(Dotl)

    # pressao = 600
    # D_casco = 60
    # li = 14 + 1/2
    # lo = 23

    # D_casco = D_casco * POL2M
    # li = li * POL2M
    # lo = lo * POL2M
    
    # Dc = 1

    # cursor = conect_sqlite(DB_CONSTANTS_DIR)
    # sql_linha = f"SELECT d_bocal FROM Diametro_bocal WHERE D_casco_min <= {Dc} AND D_casco_max >= {Dc} "
    # linha = filtro_sqlite(cursor, sql_linha, True)

    # print(linha[0])

    # angulo_tubo = 30
    # Res = 100
